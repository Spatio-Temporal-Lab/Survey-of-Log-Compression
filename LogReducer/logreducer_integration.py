#!/usr/bin/env python3
"""LogReducer 集成测试：一键跑 datasets/txt 下所有数据集并导出 Excel 报告。"""

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable, List

import importlib


def _ensure_openpyxl():
    try:
        return importlib.import_module("openpyxl")
    except ModuleNotFoundError:
        print("[INFO] openpyxl not found, installing ...")
        subprocess.run("python3 -m pip install openpyxl", shell=True, check=True)
        return importlib.import_module("openpyxl")


openpyxl = _ensure_openpyxl()


def ensure_python_packages(packages: list[str]) -> None:
    missing: list[str] = []
    for pkg in packages:
        try:
            importlib.import_module(pkg)
        except ModuleNotFoundError:
            missing.append(pkg)
    if missing:
        pkg_str = " ".join(missing)
        print(f"[INFO] installing missing packages: {pkg_str}")
        subprocess.run(f"python3 -m pip install {pkg_str}", shell=True, check=True)
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
LOGREDUCER_DIR = BASE_DIR / "LogReducer"
DEFAULT_DATASET_DIR = BASE_DIR / "datasets" / "txt"
DEFAULT_TMP_DIR = BASE_DIR / "tmp" / "logreducer_integration"
DEFAULT_OUT_DIR = BASE_DIR / "out" / "logreducer_integration"
DEFAULT_EXCEL_PATH = BASE_DIR / "results" / "logreducer_integration_results.xlsx"


@dataclass
class DatasetItem:
    name: str
    path: Path
    is_dir: bool


@dataclass
class TestResult:
    dataset: str
    original_size_mb: float
    compressed_size_mb: float | None
    compression_ratio: float | None
    compress_time_s: float | None
    decompress_time_s: float | None
    status: str
    note: str


class CommandError(RuntimeError):
    pass


def run_cmd(cmd: str, cwd: Path, timeout: int = 7200) -> None:
    print(f"    CMD: {cmd}")
    result = subprocess.run(
        cmd,
        shell=True,
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )
    output = result.stdout.decode("utf-8", errors="replace")
    if output.strip():
        lines = output.strip().splitlines()
        for line in lines[-12:]:
            print(f"    | {line}")
    if result.returncode != 0:
        raise CommandError(f"command failed with rc={result.returncode}")


def size_bytes(path: Path) -> int:
    if path.is_file():
        return path.stat().st_size
    total = 0
    for root, _dirs, files in os.walk(path):
        for name in files:
            f = Path(root) / name
            try:
                total += f.stat().st_size
            except OSError:
                continue
    return total


def discover_datasets(dataset_dir: Path) -> List[DatasetItem]:
    items: dict[str, DatasetItem] = {}
    for entry in sorted(dataset_dir.iterdir()):
        if entry.is_file() and entry.suffix == ".log":
            items[entry.stem] = DatasetItem(name=entry.stem, path=entry, is_dir=False)
        elif entry.is_dir() and entry.name not in items:
            items[entry.name] = DatasetItem(name=entry.name, path=entry, is_dir=True)
    return [items[k] for k in sorted(items)]


def merge_logs(src_dir: Path, dst_file: Path) -> None:
    log_files = sorted([p for p in src_dir.iterdir() if p.is_file() and p.suffix == ".log"])
    if not log_files:
        log_files = sorted([p for p in src_dir.iterdir() if p.is_file()])

    dst_file.parent.mkdir(parents=True, exist_ok=True)
    with dst_file.open("wb") as out_f:
        for src in log_files:
            with src.open("rb") as in_f:
                shutil.copyfileobj(in_f, out_f)


def run_logreducer_for_dataset(log_file: Path, dataset_name: str, out_dir: Path) -> tuple[float, float, int]:
    compress_dir = out_dir / dataset_name / "compressed"
    restore_dir = out_dir / dataset_name / "restored"
    template_dir = out_dir / dataset_name / "template"
    restore_file = restore_dir / f"{dataset_name}.log"

    for d in (compress_dir, restore_dir, template_dir):
        if d.exists():
            shutil.rmtree(d)
        d.mkdir(parents=True, exist_ok=True)

    train_cmd = f'python3 training.py -I "{log_file}" -T "{template_dir}" -L 5'
    run_cmd(train_cmd, LOGREDUCER_DIR)

    compress_cmd = f'python3 LogReducer.py -I "{log_file}" -T "{template_dir}" -O "{compress_dir}"'
    t0 = time.time()
    run_cmd(compress_cmd, LOGREDUCER_DIR)
    compress_time = time.time() - t0

    compressed_size = size_bytes(compress_dir)
    if compressed_size <= 0:
        raise CommandError("compressed output is empty")

    decompress_cmd = f'python3 LogRestore.py -I "{compress_dir}" -O "{restore_file}" -T "{template_dir}"'
    t1 = time.time()
    run_cmd(decompress_cmd, LOGREDUCER_DIR)
    decompress_time = time.time() - t1

    return compress_time, decompress_time, compressed_size


def export_excel(results: Iterable[TestResult], excel_path: Path) -> None:
    rows = list(results)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LogReducer_Results"

    headers = [
        "dataset",
        "original_size_mb",
        "compressed_size_mb",
        "compression_ratio",
        "compress_time_s",
        "decompress_time_s",
        "status",
        "note",
    ]
    ws.append(headers)

    for item in rows:
        data = asdict(item)
        ws.append([data[h] for h in headers])

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    end_cell = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
    table = Table(displayName="LogReducerMetrics", ref=f"A1:{end_cell}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    wb.save(excel_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="LogReducer 一键集成测试")
    parser.add_argument("--dataset-dir", type=Path, default=DEFAULT_DATASET_DIR)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--tmp-dir", type=Path, default=DEFAULT_TMP_DIR)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--dataset", type=str, default=None, help="仅运行指定数据集（目录名或 .log 文件名）")
    args = parser.parse_args()

    if not args.dataset_dir.exists():
        raise FileNotFoundError(f"dataset dir not found: {args.dataset_dir}")

    ensure_python_packages(["pandas", "numpy", "six"])

    datasets = discover_datasets(args.dataset_dir)
    if args.dataset:
        key = args.dataset[:-4] if args.dataset.endswith(".log") else args.dataset
        datasets = [d for d in datasets if d.name == key]

    if not datasets:
        raise RuntimeError("no datasets found to run")

    print(f"[INFO] total datasets: {len(datasets)}")
    print(f"[INFO] datasets: {[d.name for d in datasets]}")

    results: list[TestResult] = []
    args.tmp_dir.mkdir(parents=True, exist_ok=True)
    args.out_dir.mkdir(parents=True, exist_ok=True)

    mb = 1024 * 1024
    for ds in datasets:
        print(f"\n{'=' * 70}\n[DATASET] {ds.name}")
        if ds.is_dir:
            merged = args.tmp_dir / f"{ds.name}.log"
            if merged.exists():
                merged.unlink()
            merge_logs(ds.path, merged)
            log_file = merged
        else:
            log_file = ds.path

        original_size = size_bytes(log_file)
        original_mb = original_size / mb
        print(f"  input: {log_file} ({original_mb:.3f} MB)")

        try:
            c_time, d_time, c_size = run_logreducer_for_dataset(log_file, ds.name, args.out_dir)
            ratio = (original_size / c_size) if c_size > 0 else None
            result = TestResult(
                dataset=ds.name,
                original_size_mb=round(original_mb, 4),
                compressed_size_mb=round(c_size / mb, 4),
                compression_ratio=round(ratio, 4) if ratio else None,
                compress_time_s=round(c_time, 4),
                decompress_time_s=round(d_time, 4),
                status="SUCCESS",
                note="",
            )
            print(f"  [OK] ratio={result.compression_ratio}, comp={result.compress_time_s}s, decomp={result.decompress_time_s}s")
        except Exception as exc:
            result = TestResult(
                dataset=ds.name,
                original_size_mb=round(original_mb, 4),
                compressed_size_mb=None,
                compression_ratio=None,
                compress_time_s=None,
                decompress_time_s=None,
                status="FAILED",
                note=str(exc),
            )
            print(f"  [FAIL] {exc}")

        results.append(result)

    export_excel(results, args.excel)
    print(f"\n[INFO] excel written to: {args.excel}")

    success = sum(1 for r in results if r.status == "SUCCESS")
    failed = len(results) - success
    print(f"[INFO] summary: success={success}, failed={failed}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
